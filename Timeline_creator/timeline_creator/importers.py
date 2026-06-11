# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Bulk import — CSV paste (always available) and xlsx upload (openpyxl) (D17).

Column contract (symmetric with the io export fields):

  datetime, message, endpoint, username, account_type   (required)
  end, span_id                                           (optional; a row with
                                                          ``end`` is a span)

Validation is ALL-OR-NOTHING on STRUCTURAL errors, aggregated by row: every
row is checked through the :class:`Event` model, all failures are collected,
and the batch is committed only if every row is valid. Unknown-but-valid
endpoints/usernames are NOT errors — they are returned for auto-adding (D18).

Import timestamps may be naive; they are assumed UTC by default (the DFIR norm,
matching the "enter in UTC" entry default), then validated as aware by the
model. Pass ``assume_utc=False`` to require explicit offsets.
"""

from __future__ import annotations

import csv
import io as _io
from dataclasses import dataclass, field
from datetime import datetime, timezone

from pydantic import ValidationError

from .models import AccountType, Event

REQUIRED_COLUMNS = ("datetime", "message", "endpoint", "username", "account_type")
OPTIONAL_COLUMNS = ("end", "span_id")

# Friendly header aliases -> canonical column name.
_HEADER_ALIASES = {
    "date": "datetime",
    "timestamp": "datetime",
    "description": "message",
    "user": "username",
    "host": "endpoint",
    "hostname": "endpoint",
    "end_datetime": "end",
    "endtime": "end",
}

# ISO first; a couple of unambiguous space-separated fallbacks. Slash formats are
# intentionally rejected (dd/mm vs mm/dd is ambiguous -> ask for ISO).
_FALLBACK_FORMATS = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S")


@dataclass
class RowError:
    row: int  # 1-based data row number (header is row 0)
    message: str


@dataclass
class ImportResult:
    events: list[Event] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)
    discovered_endpoints: list[str] = field(default_factory=list)
    discovered_users: dict[str, AccountType] = field(default_factory=dict)

    @property
    def ok(self) -> bool:
        return not self.errors


class ImportError_(Exception):
    """Raised for batch-level structural problems (e.g. missing columns)."""


def _normalise_header(name: str) -> str:
    key = (name or "").strip().casefold().replace(" ", "_")
    return _HEADER_ALIASES.get(key, key)


def _parse_dt(value, assume_utc: bool) -> datetime:
    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value).strip()
        if not text:
            raise ValueError("empty datetime")
        dt = None
        try:
            dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            for fmt in _FALLBACK_FORMATS:
                try:
                    dt = datetime.strptime(text, fmt)
                    break
                except ValueError:
                    continue
        if dt is None:
            raise ValueError(
                f"unparseable datetime {value!r}; use ISO 8601 "
                "(e.g. 2025-01-01T09:00:00+00:00)"
            )
    if dt.tzinfo is None:
        if not assume_utc:
            raise ValueError(f"naive datetime {value!r}; provide an explicit offset")
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def _row_to_event(row: dict, assume_utc: bool) -> Event:
    """Build one Event from a column->value dict. Raises on any problem."""
    missing = [c for c in REQUIRED_COLUMNS if not str(row.get(c, "")).strip()]
    if missing:
        raise ValueError(f"missing required value(s): {', '.join(missing)}")

    dt = _parse_dt(row["datetime"], assume_utc)
    end_raw = row.get("end")
    end = _parse_dt(end_raw, assume_utc) if end_raw not in (None, "") else None
    account_type = AccountType(str(row["account_type"]).strip())
    span_id = (str(row["span_id"]).strip() or None) if row.get("span_id") else None

    return Event(
        datetime=dt,
        end=end,
        message=str(row["message"]).strip(),
        endpoint=str(row["endpoint"]).strip(),
        username=str(row["username"]).strip(),
        account_type=account_type,
        span_id=span_id,
    )


def _rows_to_result(header: list[str], rows: list[list], assume_utc: bool) -> ImportResult:
    """Validate a header + raw rows into an ImportResult (all-or-nothing)."""
    columns = [_normalise_header(h) for h in header]
    missing_cols = [c for c in REQUIRED_COLUMNS if c not in columns]
    if missing_cols:
        raise ImportError_(
            f"missing required column(s): {', '.join(missing_cols)}. "
            f"Expected header: {', '.join(REQUIRED_COLUMNS)} (+ optional end, span_id)."
        )

    result = ImportResult()
    for data_index, raw in enumerate(rows, start=1):
        row = {columns[i]: raw[i] for i in range(min(len(columns), len(raw)))}
        if not any(str(v).strip() for v in row.values()):
            continue  # skip wholly blank lines
        try:
            event = _row_to_event(row, assume_utc)
        except ValidationError as exc:
            messages = "; ".join(
                f"{'.'.join(str(p) for p in e['loc'])}: {e['msg']}" for e in exc.errors()
            )
            result.errors.append(RowError(data_index, messages))
        except (ValueError, KeyError) as exc:
            result.errors.append(RowError(data_index, str(exc)))
        else:
            result.events.append(event)

    if result.ok:
        for event in result.events:
            if event.endpoint not in result.discovered_endpoints:
                result.discovered_endpoints.append(event.endpoint)
            result.discovered_users.setdefault(event.username, event.account_type)
    else:
        result.events = []  # all-or-nothing: do not commit a partial batch
    return result


def parse_csv(text: str, *, assume_utc: bool = True) -> ImportResult:
    """Parse pasted CSV (with a header row) into an ImportResult."""
    reader = list(csv.reader(_io.StringIO(text)))
    rows = [r for r in reader if any(str(c).strip() for c in r)]
    if not rows:
        raise ImportError_("no data found (expected a header row + at least one row).")
    return _rows_to_result(rows[0], rows[1:], assume_utc)


def parse_xlsx(data: bytes, *, assume_utc: bool = True) -> ImportResult:
    """Parse an uploaded .xlsx (first sheet, header row) into an ImportResult.

    Requires openpyxl (confirmed available in the target env). openpyxl returns
    real datetime objects for Excel date cells, which :func:`_parse_dt` accepts.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - env dependent
        raise ImportError_(
            "xlsx import needs openpyxl, which is not installed in this environment."
        ) from exc

    workbook = load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active
    rows = [list(r) for r in sheet.iter_rows(values_only=True)]
    rows = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
    if not rows:
        raise ImportError_("no data found in the first worksheet.")
    header = [str(c) if c is not None else "" for c in rows[0]]
    return _rows_to_result(header, rows[1:], assume_utc)
